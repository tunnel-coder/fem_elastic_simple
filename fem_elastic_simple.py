# FEM_elastic_simple
import numpy as np

# создание окна
from tkinter import *
from tkinter.ttk import *
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment


def stiffness_matrix_element(B, De, E, nu, xi, xj, xm, zi, zj, zm):
    dlt = xi * (zj - zm) + xj * (zm - zi) + xm * (zi - zj)
    B[0, 0] = (zm - zj) / dlt
    B[0, 2] = (zi - zm) / dlt
    B[0, 4] = (zj - zi) / dlt
    B[1, 1] = (xj - xm) / dlt
    B[1, 3] = (xm - xi) / dlt
    B[1, 5] = (xi - xj) / dlt
    B[2, 0] = B[1, 1]
    B[2, 1] = B[0, 0]
    B[2, 2] = B[1, 3]
    B[2, 3] = B[0, 2]
    B[2, 4] = B[1, 5]
    B[2, 5] = B[0, 4]
    BT = np.transpose(B)
    dlt = np.abs(dlt)
    EnuS = E / (1 + nu) * dlt / 2
    De[0, 0] = (1 - nu) / (1 - 2 * nu) * EnuS
    De[0, 1] = nu / (1 - 2 * nu) * EnuS
    De[0, 2] = 0
    De[1, 0] = De[0, 1]
    De[1, 1] = De[0, 0]
    De[1, 2] = 0
    De[2, 0] = 0
    De[2, 1] = 0
    De[2, 2] = 0.5 * EnuS
    BTD = np.dot(BT, De)
    BTDB = np.dot(BTD, B)
    kloc = BTDB
    return kloc


def stiffness_counter(B, De, E, nu, xi, xj, xm, zi, zj, zm):
    dlt = xi * (zj - zm) + xj * (zm - zi) + xm * (zi - zj)
    B[0, 0] = (zm - zj) / dlt
    B[0, 2] = (zi - zm) / dlt
    B[0, 4] = (zj - zi) / dlt
    B[1, 1] = (xj - xm) / dlt
    B[1, 3] = (xm - xi) / dlt
    B[1, 5] = (xi - xj) / dlt
    B[2, 0] = B[1, 1]
    B[2, 1] = B[0, 0]
    B[2, 2] = B[1, 3]
    B[2, 3] = B[0, 2]
    B[2, 4] = B[1, 5]
    B[2, 5] = B[0, 4]
    dlt = np.abs(dlt)
    EnuS = E / (1 + nu) * dlt / 2
    De[0, 0] = (1 - nu) / (1 - 2 * nu) * EnuS
    De[0, 1] = nu / (1 - 2 * nu) * EnuS
    De[0, 2] = 0
    De[1, 0] = De[0, 1]
    De[1, 1] = De[0, 0]
    De[1, 2] = 0
    De[2, 0] = 0
    De[2, 1] = 0
    De[2, 2] = 0.5 * EnuS
    return dlt


def stiffness_matrix_total(kglob, ki, kj, kloc, km):
    for i1 in range(0, 6):
        if 0 <= i1 <= 1:
            nk1 = ki * 2 + i1 - 2
        elif 2 <= i1 <= 3:
            nk1 = kj * 2 + i1 - 4
        elif 4 <= i1 <= 5:
            nk1 = km * 2 + i1 - 6
        for j1 in range(0, 6):
            if 0 <= j1 <= 1:
                nk2 = ki * 2 + j1 - 2
            elif 2 <= j1 <= 3:
                nk2 = kj * 2 + j1 - 4
            elif 4 <= j1 <= 5:
                nk2 = km * 2 + j1 - 6
            kglob[nk1, nk2] = kglob[nk1, nk2] + kloc[i1, j1]
    return kglob


class Application(Frame):
    def __init__(self, master):
        super(Application, self).__init__(master)
        self.grid()
        self.create_widgets()

    def create_widgets(self):
        # Elastic_modulus
        Label(self,
              text="E ="
              ).grid(row=0, column=0)
        self.E_ent = Entry(self, width=5, justify=CENTER)
        self.E_ent.grid(row=0, column=1)
        self.E_ent.insert(0, "10000")
        Label(self,
              text="kPa"
              ).grid(row=0, column=2, sticky=W)
        # Poissons_ratio
        Label(self,
              text="v ="
              ).grid(row=1, column=0)
        self.nu_ent = Entry(self, width=5, justify=CENTER)
        self.nu_ent.grid(row=1, column=1)
        self.nu_ent.insert(0, "0.3")
        Label(self,
              text=""
              ).grid(row=1, column=2, sticky=W)
        # Pressure
        Label(self,
              text="P ="
              ).grid(row=2, column=0)
        self.pLoad_ent = Entry(self, width=5, justify=CENTER)
        self.pLoad_ent.grid(row=2, column=1)
        self.pLoad_ent.insert(0, "1000")
        Label(self,
              text="kPa"
              ).grid(row=2, column=2, sticky=W)
        # Stamp width
        Label(self,
              text="b ="
              ).grid(row=3, column=0)
        self.b_ent = Entry(self, width=5, justify=CENTER)
        self.b_ent.grid(row=3, column=1)
        self.b_ent.insert(0, "2")
        Label(self,
              text="m"
              ).grid(row=3, column=2, sticky=W)
        # Object widts
        Label(self,
              text="ws ="
              ).grid(row=4, column=0)
        self.ws_ent = Entry(self, width=5, justify=CENTER)
        self.ws_ent.grid(row=4, column=1)
        self.ws_ent.insert(0, "5")
        Label(self,
              text="m"
              ).grid(row=4, column=2, sticky=W)
        # Object height
        Label(self,
              text="hs ="
              ).grid(row=5, column=0)
        self.hs_ent = Entry(self, width=5, justify=CENTER)
        self.hs_ent.grid(row=5, column=1)
        self.hs_ent.insert(0, "4")
        Label(self,
              text="m"
              ).grid(row=5, column=2, sticky=W)
        # Width_number
        Label(self,
              text="nw ="
              ).grid(row=0, column=4)
        self.nw_ent = Entry(self, width=5, justify=CENTER)
        self.nw_ent.grid(row=0, column=5)
        self.nw_ent.insert(0, "20")
        Label(self,
              text=""
              ).grid(row=0, column=6, sticky=W)
        # Height_number
        Label(self,
              text="nh ="
              ).grid(row=1, column=4)
        self.nh_ent = Entry(self, width=5, justify=CENTER)
        self.nh_ent.grid(row=1, column=5)
        self.nh_ent.insert(0, "16")
        Label(self,
              text=""
              ).grid(row=1, column=6, sticky=W)
        # Stamp_number
        Label(self,
              text="nb ="
              ).grid(row=2, column=4)
        self.nb_ent = Entry(self, width=5, justify=CENTER)
        self.nb_ent.grid(row=2, column=5)
        self.nb_ent.insert(0, "12")
        Label(self,
              text=""
              ).grid(row=2, column=6, sticky=W)
        # Btn_count
        Button(self,
               text="Решить, записать результаты в Excel",
               command=self.solver
               ).grid(row=7, column=0, columnspan=6)

    def input(self):
        E = float(self.E_ent.get())
        nu = float(self.nu_ent.get())
        p_load = float(self.pLoad_ent.get())
        b = int(self.b_ent.get())
        ws = int(self.ws_ent.get())
        hs = int(self.hs_ent.get())
        nw = int(self.nw_ent.get())
        nh = int(self.nh_ent.get())
        nb = int(self.nb_ent.get())
        return E, nu, p_load, b, ws, hs, nw, nh, nb

    def solver(self):
        E, nu, p_load, b, ws, hs, nw, nh, nb = self.input()
        # Mesh create
        b = b / 2
        # Nodes number
        n = (nw + 1) * (nh + 1)
        # Elements number
        nel = nw * nh * 2
        "n_band = 2 * (nh + 2)"
        kx = np.zeros((nh + 1, nw + 1))
        kz = np.zeros((nh + 1, nw + 1))
        xk = 0
        zk = 0
        for i in range(nb):
            kx[0, i] = xk
            xk += b / nb
        for j in range(nb, nw + 1):
            kx[0, j] = xk
            xk += (ws - b) / (nw - nb)
        for k in range(1, nh + 1):
            kx[k] = kx[0]
        for i in range(nh + 1):
            kz[i, 0] = zk
            zk += hs / nh
        for i in range(1, nw + 1):
            for j in range(1, nh + 1):
                kz[j, i] = kz[j, i - 1]
        kxt = np.transpose(kx)
        kzt = np.transpose(kz)

        # Nodes numbers
        n_nmb = np.zeros((nh + 1, nw + 1))
        number = 1
        for i in range(nw + 1):
            for j in range(nh + 1):
                n_nmb[j, i] = number
                number += 1

        # Stiffness matrix
        B = np.zeros((3, 6))
        De = np.zeros((3, 3))
        kglob = np.zeros((n * 2, n * 2))
        k = 0
        for i in range(1, nw + 1):
            for j in range(1, nh + 1):
                k = k + 1
                l = (k + 1) // 2 + i - 1
                ki = l
                kj = l + 1
                km = l + nh + 1
                # Stiffness_matrix_element
                xi = kx[j - 1, i - 1]
                xj = kx[j, i - 1]
                xm = kx[j - 1, i]
                zi = kz[j - 1, i - 1]
                zj = kz[j, i - 1]
                zm = kz[j - 1, i]
                kloc = stiffness_matrix_element(B, De, E, nu, xi, xj, xm, zi, zj, zm)
                kglob = stiffness_matrix_total(kglob, ki, kj, kloc, km)

                k = k + 1
                m = (k + 2) // 2 + i - 1
                ki = m
                kj = m + nh
                km = m + nh + 1
                # Stiffness_matrix_element
                xi = kx[j, i - 1]
                xj = kx[j - 1, i]
                xm = kx[j, i]
                zi = kz[j, i - 1]
                zj = kz[j - 1, i]
                zm = kz[j, i]
                kloc = stiffness_matrix_element(B, De, E, nu, xi, xj, xm, zi, zj, zm)
                kglob = stiffness_matrix_total(kglob, ki, kj, kloc, km)

        # Boundary conditions
        R = np.zeros(n * 2)
        u = np.zeros(n * 2)

        for i in range(1, nb + 2):
            j = ((i - 1) * nh + i) * 2
            if i == 1 or i == nb + 1:
                R[j - 1] = 0.5 * p_load * b / nb
            else:
                R[j - 1] = p_load * b / nb

        for i in range(1, nh + 2):
            j = i * 2
            u[j - 2] = 0
            kglob[j - 2, j - 2] = 10 ** 16
            j = (n - i + 1) * 2
            u[j - 2] = 0
            u[j - 1] = 0
            kglob[j - 1, j - 1] = 10 ** 16
            kglob[j - 2, j - 2] = 10 ** 16

        for i in range(nh + 1, n, nh + 1):
            j = i * 2
            u[j - 1] = 0
            kglob[j - 1, j - 1] = 10 ** 16

        # Solve SOLE
        kglob_obr = np.linalg.inv(kglob)
        u = np.dot(kglob_obr, R)

        # Get Strains and deforms
        eps_x = np.zeros(nel)
        eps_z = np.zeros(nel)
        eps_xz = np.zeros(nel)
        sigma_x = np.zeros(nel)
        sigma_z = np.zeros(nel)
        tau_xz = np.zeros(nel)
        sigma_1 = np.zeros(nel)
        sigma_3 = np.zeros(nel)

        k = 0
        for i in range(1, nw + 1):
            for j in range(1, nh + 1):
                k = k + 1
                l = (k + 1) // 2 + i - 1
                ki = l
                kj = l + 1
                km = l + nh + 1
                # Stiffness_matrix_element
                xi = kx[j - 1, i - 1]
                xj = kx[j, i - 1]
                xm = kx[j - 1, i]
                zi = kz[j - 1, i - 1]
                zj = kz[j, i - 1]
                zm = kz[j - 1, i]

                dlt = stiffness_counter(B, De, E, nu, xi, xj, xm, zi, zj, zm)

                eps_x[k - 1] = (u[2 * ki - 2] * B[0, 0] + u[2 * kj - 2] * B[0, 2] + u[2 * km - 2] * B[0, 4])
                eps_z[k - 1] = (u[2 * ki - 1] * B[1, 1] + u[2 * kj - 1] * B[1, 3] + u[2 * km - 1] * B[1, 5])
                eps_xz[k - 1] = (u[2 * ki - 2] * B[2, 0] + u[2 * kj - 2] * B[2, 2] + u[2 * km - 2] * B[2, 4]
                                 + u[2 * ki - 1] * B[2, 1] + u[2 * kj - 1] * B[2, 3] + u[2 * km - 1] * B[2, 5])

                sigma_x[k - 1] = (De[0, 0] * eps_x[k - 1] + De[0, 1] * eps_z[k - 1]) / dlt * 2
                sigma_z[k - 1] = (De[1, 0] * eps_x[k - 1] + De[1, 1] * eps_z[k - 1]) / dlt * 2
                tau_xz[k - 1] = De[2, 2] * eps_xz[k - 1] / dlt * 2

                sigma_1[k - 1] = (sigma_x[k - 1] + sigma_z[k - 1]) / 2 + np.sqrt((sigma_x[k - 1] - sigma_z[k - 1])
                                                                                 ** 2 / 4 + tau_xz[k - 1] ** 2)
                sigma_3[k - 1] = (sigma_x[k - 1] + sigma_z[k - 1]) / 2 - np.sqrt((sigma_x[k - 1] - sigma_z[k - 1])
                                                                                 ** 2 / 4 + tau_xz[k - 1] ** 2)

                k = k + 1
                m = (k + 2) // 2 + i - 1
                ki = m
                kj = m + nh
                km = m + nh + 1
                # Stiffness_matrix_element
                xi = kx[j, i - 1]
                xj = kx[j - 1, i]
                xm = kx[j, i]
                zi = kz[j, i - 1]
                zj = kz[j - 1, i]
                zm = kz[j, i]

                dlt = stiffness_counter(B, De, E, nu, xi, xj, xm, zi, zj, zm)

                eps_x[k - 1] = (u[2 * ki - 2] * B[0, 0] + u[2 * kj - 2] * B[0, 2] + u[2 * km - 2] * B[0, 4])
                eps_z[k - 1] = (u[2 * ki - 1] * B[1, 1] + u[2 * kj - 1] * B[1, 3] + u[2 * km - 1] * B[1, 5])
                eps_xz[k - 1] = (u[2 * ki - 2] * B[2, 0] + u[2 * kj - 2] * B[2, 2] + u[2 * km - 2] * B[2, 4]
                                 + u[2 * ki - 1] * B[2, 1] + u[2 * kj - 1] * B[2, 3] + u[2 * km - 1] * B[2, 5])

                sigma_x[k - 1] = (De[0, 0] * eps_x[k - 1] + De[0, 1] * eps_z[k - 1]) / dlt * 2
                sigma_z[k - 1] = (De[1, 0] * eps_x[k - 1] + De[1, 1] * eps_z[k - 1]) / dlt * 2
                tau_xz[k - 1] = De[2, 2] * eps_xz[k - 1] / dlt * 2

                sigma_1[k - 1] = (sigma_x[k - 1] + sigma_z[k - 1]) / 2 + np.sqrt((sigma_x[k - 1] - sigma_z[k - 1])
                                                                                 ** 2 / 4 + tau_xz[k - 1] ** 2)
                sigma_3[k - 1] = (sigma_x[k - 1] + sigma_z[k - 1]) / 2 - np.sqrt((sigma_x[k - 1] - sigma_z[k - 1])
                                                                                 ** 2 / 4 + tau_xz[k - 1] ** 2)

        u_new = np.zeros(n)
        for i, j in zip(range(0, n * 2, 2), range(n + 1)):
            u_new[j] = u[i]
        u_new = np.transpose(np.reshape(u_new, (nw + 1, nh + 1)))
        kx_new = kx + u_new

        w_new = np.zeros(n)
        for i, j in zip(range(1, n * 2 + 1, 2), range(n + 1)):
            w_new[j] = u[i]
        w_new = np.transpose(np.reshape(w_new, (nw + 1, nh + 1)))
        kz_new = kz + w_new

        kxt_new = np.transpose(kx_new)
        kzt_new = np.transpose(kz_new)

        # Graphics
        fig = plt.figure()
        ax = fig.add_subplot()
        ax.set_title('Mesh')
        ax.plot(kx, -kz, color='blue', linewidth=0.4)
        ax.plot(kxt, -kzt, color='blue', linewidth=0.4)
        for i in range(nh):
            for j in range(nw):
                ax.plot([kx[i, j], kx[i + 1, j + 1]], [-kz[i + 1, j + 1], -kz[i, j]], color='blue', linewidth=0.4)

        ax.plot(kx_new, -kz_new, color='red', linewidth=0.4)
        ax.plot(kxt_new, -kzt_new, color='red', linewidth=0.4)

        for i in range(nh + 1):
            for j in range(nw + 1):
                ax.text(kx[i, j], -kz[i, j], n_nmb[i, j].astype(int), fontsize=5, fontstyle='italic')
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.yaxis.set_major_locator(ticker.MultipleLocator(1))
        plt.gca().set_aspect('equal', adjustable='box')
        plt.show()

        # Table_creation
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_center = Alignment(horizontal='center',
                                 vertical='bottom',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active
        ws.sheet_view.zoomScale = 85

        # название страницы
        ws.title = 'Матрица жесткости'

        for i in range(n * 2):
            for j in range(n * 2):
                ws.cell(row=i + 1, column=j + 1).value = kglob[i, j]
                ws.cell(row=i + 1, column=j + 1).alignment = align_center
                ws.cell(row=i + 1, column=j + 1).border = border
                ws.cell(row=i + 1, column=j + 1).number_format = '0.00'

        # второй лист
        ws1 = wb.create_sheet("Элементы")
        ws1.sheet_view.zoomScale = 85
        column = ['№ эл-та', 'sigma_x', 'sigma_z', 'tau_xz', 'sigma_1', 'sigma_3', 'eps_x', 'eps_z', 'eps_xz']
        for i, value in enumerate(column):
            ws1.cell(row=1, column=i + 1).value = value
            ws1.cell(row=1, column=i + 1).fill = fill
            ws1.cell(row=1, column=i + 1).alignment = align_center
            ws1.cell(row=1, column=i + 1).border = border
        for i in range(nel):
            ws1.cell(row=i + 2, column=1).value = i + 1
            ws1.cell(row=i + 2, column=1).alignment = align_center
            ws1.cell(row=i + 2, column=1).border = border
            ws1.cell(row=i + 2, column=2).value = sigma_x[i]
            ws1.cell(row=i + 2, column=2).alignment = align_center
            ws1.cell(row=i + 2, column=2).border = border
            ws1.cell(row=i + 2, column=2).number_format = '0.000'
            ws1.cell(row=i + 2, column=3).value = sigma_z[i]
            ws1.cell(row=i + 2, column=3).alignment = align_center
            ws1.cell(row=i + 2, column=3).border = border
            ws1.cell(row=i + 2, column=3).number_format = '0.000'
            ws1.cell(row=i + 2, column=4).value = tau_xz[i]
            ws1.cell(row=i + 2, column=4).alignment = align_center
            ws1.cell(row=i + 2, column=4).border = border
            ws1.cell(row=i + 2, column=4).number_format = '0.000'
            ws1.cell(row=i + 2, column=5).value = sigma_1[i]
            ws1.cell(row=i + 2, column=5).alignment = align_center
            ws1.cell(row=i + 2, column=5).border = border
            ws1.cell(row=i + 2, column=5).number_format = '0.000'
            ws1.cell(row=i + 2, column=6).value = sigma_3[i]
            ws1.cell(row=i + 2, column=6).alignment = align_center
            ws1.cell(row=i + 2, column=6).border = border
            ws1.cell(row=i + 2, column=6).number_format = '0.000'
            ws1.cell(row=i + 2, column=7).value = eps_x[i]
            ws1.cell(row=i + 2, column=7).alignment = align_center
            ws1.cell(row=i + 2, column=7).border = border
            ws1.cell(row=i + 2, column=7).number_format = '0.000'
            ws1.cell(row=i + 2, column=8).value = eps_z[i]
            ws1.cell(row=i + 2, column=8).alignment = align_center
            ws1.cell(row=i + 2, column=8).border = border
            ws1.cell(row=i + 2, column=8).number_format = '0.000'
            ws1.cell(row=i + 2, column=9).value = eps_xz[i]
            ws1.cell(row=i + 2, column=9).alignment = align_center
            ws1.cell(row=i + 2, column=9).border = border
            ws1.cell(row=i + 2, column=9).number_format = '0.000'

        wb.save("Results.xlsx")


main_window = Tk()
main_window.title("FEM elastic simple")
main_window.geometry("260x170")
main_window.resizable(0, 0)
main_window.attributes("-toolwindow", 0)
app = Application(main_window)
main_window.mainloop()
